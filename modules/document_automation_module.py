import os
import json
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import jinja2
from utils import get_logger, ModuleValidator, prompt_loader
from config import ProcureMateSettings

logger = get_logger(__name__)

class DocumentAutomationModule:
    """문서 자동화 모듈"""
    
    def __init__(self):
        self.output_path = Path(ProcureMateSettings.DOCUMENT_OUTPUT_PATH)
        self.template_path = Path(ProcureMateSettings.TEMPLATE_PATH)
        self.validator = ModuleValidator("DocumentAutomationModule")
        
        self._ensure_directories()
        self._setup_jinja_env()
        
        logger.info("DocumentAutomationModule 초기화")
    
    def _ensure_directories(self):
        """필요한 디렉토리 생성"""
        self.output_path.mkdir(exist_ok=True)
        self.template_path.mkdir(exist_ok=True)
        
        # 기본 템플릿 생성
        self._create_default_templates()
    
    def _setup_jinja_env(self):
        """Jinja2 환경 설정"""
        self.jinja_env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(str(self.template_path)),
            autoescape=jinja2.select_autoescape(['html', 'xml'])
        )
    
    def _create_default_templates(self):
        """기본 HTML 템플릿 생성"""
        template_file = self.template_path / "procurement_request.html"
        if not template_file.exists():
            # prompt_loader를 사용하여 템플릿 로드
            template_content = prompt_loader.load_template("procurement_request.html")
            if template_content:
                template_file.write_text(template_content, encoding='utf-8')
                logger.info("기본 조달 요청서 템플릿 생성")
    
    def generate_procurement_pdf(self, data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """조달 요청서 PDF 생성 (reportlab 사용)"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"procurement_request_{timestamp}.pdf"
        
        output_file = self.output_path / filename
        doc = SimpleDocTemplate(str(output_file), pagesize=A4)
        styles = getSampleStyleSheet()
        
        # 스타일 정의
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1  # 중앙 정렬
        )
        
        # 문서 요소들
        story = []
        
        # 제목
        story.append(Paragraph("조달 요청서", title_style))
        story.append(Paragraph(f"작성일: {data.get('created_date', datetime.now().strftime('%Y-%m-%d'))}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # 기본 정보
        story.append(Paragraph("<b>요청 개요</b>", styles['Heading2']))
        story.append(Paragraph(f"요청자: {data.get('requester', '')}", styles['Normal']))
        story.append(Paragraph(f"요청 내용: {data.get('request_description', '')}", styles['Normal']))
        story.append(Paragraph(f"긴급도: {data.get('urgency', '')}", styles['Normal']))
        story.append(Paragraph(f"예산 범위: {data.get('budget_range', '')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # 물품 목록 테이블
        if data.get('items'):
            story.append(Paragraph("<b>요청 물품 목록</b>", styles['Heading2']))
            
            table_data = [['물품명', '수량', '예상 단가', '예상 총액', '비고']]
            for item in data['items']:
                table_data.append([
                    item.get('name', ''),
                    item.get('quantity', ''),
                    item.get('unit_price', ''),
                    item.get('total_price', ''),
                    item.get('notes', '')
                ])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # 추천 업체 정보
        if data.get('recommendations'):
            story.append(Paragraph("<b>추천 업체 정보</b>", styles['Heading2']))
            for rec in data['recommendations']:
                story.append(Paragraph(
                    f"<b>{rec.get('platform', '')}</b> - {rec.get('vendor', '')}<br/>"
                    f"상품명: {rec.get('name', '')}<br/>"
                    f"가격: {rec.get('price', 0):,}원<br/>"
                    f"평점: {rec.get('rating', 0)} (리뷰 {rec.get('review_count', 0)}개)",
                    styles['Normal']
                ))
                story.append(Spacer(1, 10))
        
        # 푸터
        story.append(Spacer(1, 50))
        story.append(Paragraph("ProcureMate 시스템에서 자동 생성됨", styles['Normal']))
        
        # PDF 생성
        doc.build(story)
        
        logger.info(f"PDF 생성 완료: {output_file}")
        return str(output_file)


    def generate_procurement_excel(self, data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """조달 요청서 Excel 생성"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"procurement_request_{timestamp}.xlsx"
            
            # 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "조달 요청서"
            
            # 스타일 정의
            header_font = Font(bold=True, size=14)
            title_font = Font(bold=True, size=16)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 제목
            ws.merge_cells('A1:E1')
            ws['A1'] = "조달 요청서"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # 기본 정보
            row = 3
            ws[f'A{row}'] = "작성일:"
            ws[f'B{row}'] = data.get('created_date', datetime.now().strftime('%Y-%m-%d'))
            
            row += 1
            ws[f'A{row}'] = "요청자:"
            ws[f'B{row}'] = data.get('requester', '')
            
            row += 1
            ws[f'A{row}'] = "요청 내용:"
            ws[f'B{row}'] = data.get('request_description', '')
            
            row += 1
            ws[f'A{row}'] = "긴급도:"
            ws[f'B{row}'] = data.get('urgency', '')
            
            row += 1
            ws[f'A{row}'] = "예산 범위:"
            ws[f'B{row}'] = data.get('budget_range', '')
            
            # 물품 목록 테이블
            row += 2
            headers = ['물품명', '수량', '예상 단가', '예상 총액', '비고']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.border = border
            
            # 물품 데이터
            for item in data.get('items', []):
                row += 1
                ws.cell(row=row, column=1, value=item.get('name', '')).border = border
                ws.cell(row=row, column=2, value=item.get('quantity', '')).border = border
                ws.cell(row=row, column=3, value=item.get('unit_price', '')).border = border
                ws.cell(row=row, column=4, value=item.get('total_price', '')).border = border
                ws.cell(row=row, column=5, value=item.get('notes', '')).border = border
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30
            
            # 파일 저장
            output_file = self.output_path / filename
            wb.save(str(output_file))
            
            logger.info(f"Excel 생성 완료: {output_file}")
            return str(output_file)
            
        except Exception as e:
            logger.error(f"Excel 생성 중 오류: {str(e)}")
            return ""
    
    def generate_comparison_report(self, search_results: Dict[str, List[Dict]], filename: Optional[str] = None) -> str:
        """가격 비교 보고서 생성"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"price_comparison_{timestamp}.xlsx"
            
            wb = openpyxl.Workbook()
            
            for platform, products in search_results.items():
                if not products:
                    continue
                    
                ws = wb.create_sheet(title=platform.upper())
                
                # 헤더
                headers = ['상품명', '가격', '업체', '평점', '리뷰수', 'URL']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # 데이터
                for row, product in enumerate(products, 2):
                    ws.cell(row=row, column=1, value=product.get('name', ''))
                    ws.cell(row=row, column=2, value=product.get('price', 0))
                    ws.cell(row=row, column=3, value=product.get('vendor', ''))
                    ws.cell(row=row, column=4, value=product.get('rating', 0))
                    ws.cell(row=row, column=5, value=product.get('review_count', 0))
                    ws.cell(row=row, column=6, value=product.get('product_url', ''))
                
                # 열 너비 조정
                for col in range(1, 7):
                    ws.column_dimensions[chr(64 + col)].width = 20
            
            # 기본 시트 삭제
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            output_file = self.output_path / filename
            wb.save(str(output_file))
            
            logger.info(f"가격 비교 보고서 생성 완료: {output_file}")
            return str(output_file)
            
        except Exception as e:
            logger.error(f"가격 비교 보고서 생성 중 오류: {str(e)}")
            return ""
    
    def create_procurement_document_package(self, analysis: Dict, search_results: Dict, recommendations: str) -> Dict[str, str]:
        """조달 문서 패키지 생성"""
        logger.info("조달 문서 패키지 생성 시작")
        
        # 문서 데이터 준비
        document_data = {
            'created_date': datetime.now().strftime('%Y-%m-%d'),
            'requester': 'ProcureMate 시스템',
            'request_description': f"AI 분석 기반 조달 요청",
            'urgency': analysis.get('urgency', '보통'),
            'budget_range': analysis.get('budget_range', '미정'),
            'items': []
        }
        
        # 물품 목록 구성
        items = analysis.get('items', [])
        quantities = analysis.get('quantities', [])
        
        for i, item in enumerate(items):
            item_data = {
                'name': item,
                'quantity': quantities[i] if i < len(quantities) else '1',
                'unit_price': '견적 필요',
                'total_price': '견적 필요',
                'notes': ''
            }
            document_data['items'].append(item_data)
        
        # 추천 정보 추가
        if search_results:
            document_data['recommendations'] = []
            for platform, products in search_results.items():
                for product in products[:3]:  # 상위 3개만
                    rec = {
                        'platform': platform,
                        'vendor': product.get('vendor', ''),
                        'name': product.get('name', ''),
                        'price': product.get('price', 0),
                        'rating': product.get('rating', 0),
                        'review_count': product.get('review_count', 0)
                    }
                    document_data['recommendations'].append(rec)
        
        # 문서 생성
        results = {}
        
        # PDF 생성
        pdf_file = self.generate_procurement_pdf(document_data)
        if pdf_file:
            results['pdf'] = pdf_file
        
        # Excel 생성
        excel_file = self.generate_procurement_excel(document_data)
        if excel_file:
            results['excel'] = excel_file
        
        # 가격 비교 보고서
        if search_results:
            comparison_file = self.generate_comparison_report(search_results)
            if comparison_file:
                results['comparison'] = comparison_file
        
        logger.info(f"문서 패키지 생성 완료: {list(results.keys())}")
        return results
    
    def run_validation_tests(self) -> bool:
        """모듈 검증 테스트"""
        test_data = {
            'created_date': '2025-05-26',
            'requester': '테스트 사용자',
            'request_description': '테스트 조달 요청',
            'urgency': '보통',
            'budget_range': '100만원',
            'items': [
                {
                    'name': '테스트 상품',
                    'quantity': '1',
                    'unit_price': '50000',
                    'total_price': '50000',
                    'notes': '테스트용'
                }
            ]
        }
        
        test_cases = [
            {
                "input": {"data": test_data},
                "expected": None
            }
        ]
        
        def test_pdf_generation(data: Dict):
            result = self.generate_procurement_pdf(data)
            return len(result) > 0 and result.endswith('.pdf')
        
        def test_excel_generation(data: Dict):
            result = self.generate_procurement_excel(data)
            return len(result) > 0 and result.endswith('.xlsx')
        
        pdf_result = self.validator.validate_function(test_pdf_generation, test_cases)
        excel_result = self.validator.validate_function(test_excel_generation, test_cases)
        
        debug_info = self.validator.debug_module_state(self)
        summary = self.validator.get_test_summary()
        
        logger.info(f"DocumentAutomationModule validation summary: {summary}")
        
        return pdf_result and excel_result

# 디버그 실행
if __name__ == "__main__":
    logger.info("DocumentAutomationModule 디버그 모드 시작")
    
    doc_module = DocumentAutomationModule()
    
    # 테스트 데이터
    test_analysis = {
        'items': ['사무용 의자', '책상'],
        'quantities': ['2', '1'],
        'urgency': '보통',
        'budget_range': '50만원'
    }
    
    test_search_results = {
        'coupang': [
            {
                'name': '사무용 의자 프리미엄',
                'price': 150000,
                'vendor': '가구업체A',
                'rating': 4.5,
                'review_count': 200,
                'product_url': 'http://example.com'
            }
        ]
    }
    
    # 문서 패키지 생성 테스트
    package = doc_module.create_procurement_document_package(
        test_analysis, 
        test_search_results, 
        "AI 추천 내용"
    )
    
    logger.info(f"생성된 문서: {package}")
    
    # 검증 테스트
    if doc_module.run_validation_tests():
        logger.info("모든 검증 테스트 통과")
    else:
        logger.error("검증 테스트 실패")
    
    logger.info("DocumentAutomationModule 디버그 완료")
